from flask import Flask, request, render_template, redirect, url_for, flash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from models import session as db_session, Order
import os

app = Flask(__name__)
app.secret_key = 'crm_secret_key'  # нужен для flash-сообщений

# Настройки загрузки
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Создаём папку, если не существует
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Разрешённые форматы файлов
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Главная страница для загрузки Excel
@app.route('/upload', methods=['GET', 'POST'])
def upload_excel():
    if request.method == 'POST':
        file = request.files.get('file')
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            # Читаем Excel-файл
            wb = load_workbook(filepath)
            ws = wb.active

            for row in ws.iter_rows(min_row=2):  # начинаем со 2-й строки
                track_code = row[0].value
                status = row[1].value
                flight = row[2].value

                if not track_code or status is None:
                    continue

                # Обновляем или создаём запись
                order = db_session.query(Order).filter_by(track_code=track_code).first()
                if order:
                    order.status = status
                    order.flight = flight
                else:
                    new_order = Order(track_code=track_code, status=status, flight=flight)
                    db_session.add(new_order)

            db_session.commit()
            flash("Файл успешно загружен и обработан ✅", "success")
            return redirect(url_for('upload_excel'))
        else:
            flash("Ошибка: загрузите файл .xlsx", "error")

    return render_template('upload.html')


if __name__ == '__main__':
    app.run(debug=True, host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
